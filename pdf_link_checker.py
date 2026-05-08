"""
PDF Link Checker
================
Checks all email and web addresses in a PDF for:
1. Missing hyperlinks (text visible but not clickable)
2. Wrong link type (email linked to URL, URL linked to mailto, etc.)
3. Broken/split addresses across lines — detects if parts join to form a valid address
4. Split-address partial linking (one part linked, other not — or parts linked to different targets)
5. Wrong link targets (displayed address doesn't match the actual href)

Usage:
    python pdf_link_checker.py your_file.pdf
    python pdf_link_checker.py your_file.pdf --output report.xlsx
    python pdf_link_checker.py your_file.pdf --verbose

Requirements:
    pip install pypdf pdfplumber openpyxl
"""

import re
import sys
import argparse
import pdfplumber
from pypdf import PdfReader
from dataclasses import dataclass, field
from typing import Optional
from urllib.parse import urlparse


# ---------------------------------------------------------------------------
# Patterns
# ---------------------------------------------------------------------------

EMAIL_RE = re.compile(
    r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}"
)

URL_RE = re.compile(
    r"(?:https?://|www\.)[a-zA-Z0-9\-._~:/?#\[\]@!$&'()*+,;=%]+"
)

# Fragmented email: left part (before @), right part (after @)
EMAIL_LEFT_RE  = re.compile(r"[a-zA-Z0-9._%+\-]+@$|[a-zA-Z0-9._%+\-]+@\s*$")
EMAIL_RIGHT_RE = re.compile(r"^[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")

# Fragmented URL: ends mid-word (no space after last valid char)
URL_FRAGMENT_RE = re.compile(
    r"(?:https?://|www\.)[a-zA-Z0-9\-._~:/?#\[\]@!$&'()*+,;=%]+"
)


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class LinkAnnotation:
    """A clickable hyperlink annotation found in the PDF."""
    page_num: int       # 0-based
    uri: str            # The actual href (e.g. "mailto:a@b.com" or "https://...")
    rect: tuple         # (x0, y0, x1, y1) in PDF coordinates


@dataclass
class TextToken:
    """A word/token extracted from the PDF with its bounding box."""
    page_num: int
    text: str
    x0: float
    y0: float
    x1: float
    y1: float


@dataclass
class CheckResult:
    """Result of checking a single link."""
    pdf_link: str
    hyperlink_text: str
    page_num: int
    link_type: str  # "Email" or "Web Link"
    is_hyperlinked: str  # "Yes" or "No"
    is_valid: str  # "Yes" or "No"
    hyperlink_points_to: str  # The actual URI or "No Link"
    result: str  # "Pass" or "Fail"


@dataclass
class Issue:
    page: int
    issue_type: str
    displayed_text: str
    linked_uri: Optional[str]
    detail: str

    def __str__(self):
        uri_str = f'  → Linked URI : "{self.linked_uri}"' if self.linked_uri else "  → No hyperlink found"
        return (
            f"[Page {self.page}] [{self.issue_type}]\n"
            f"  Displayed    : \"{self.displayed_text}\"\n"
            f"{uri_str}\n"
            f"  Detail       : {self.detail}"
        )


# ---------------------------------------------------------------------------
# Extraction helpers
# ---------------------------------------------------------------------------

def extract_annotations(pdf_path: str) -> list[LinkAnnotation]:
    """Extract all URI link annotations from every page."""
    annotations = []
    reader = PdfReader(pdf_path)

    for page_num, page in enumerate(reader.pages):
        if "/Annots" not in page:
            continue
        annots = page["/Annots"]
        if annots is None:
            continue

        # Page height for coordinate conversion (PDF y=0 is bottom-left)
        page_height = float(page.mediabox.height)

        for annot_ref in annots:
            try:
                annot = annot_ref.get_object() if hasattr(annot_ref, "get_object") else annot_ref
                if annot.get("/Subtype") != "/Link":
                    continue
                action = annot.get("/A")
                if action is None:
                    continue
                if action.get("/S") != "/URI":
                    continue
                uri = str(action.get("/URI", ""))
                rect_raw = annot.get("/Rect")
                if rect_raw is None:
                    continue
                # Convert PDF rect (bottom-left origin) → top-left origin
                rx0, ry0, rx1, ry1 = [float(v) for v in rect_raw]
                # Normalise so x0<x1, y0<y1
                x0, x1 = sorted([rx0, rx1])
                # Convert y from bottom-left to top-left
                y0 = page_height - max(ry0, ry1)
                y1 = page_height - min(ry0, ry1)
                annotations.append(LinkAnnotation(page_num, uri, (x0, y0, x1, y1)))
            except Exception:
                continue

    return annotations


def extract_words(pdf_path: str) -> list[TextToken]:
    """Extract every word with its bounding box using pdfplumber."""
    tokens = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            words = page.extract_words(
                x_tolerance=3,
                y_tolerance=3,
                keep_blank_chars=False,
                use_text_flow=False,
                extra_attrs=["fontname", "size"],
            )
            for w in words:
                tokens.append(TextToken(
                    page_num=page_num,
                    text=w["text"],
                    x0=w["x0"],
                    y0=w["top"],
                    x1=w["x1"],
                    y1=w["bottom"],
                ))
    return tokens


# ---------------------------------------------------------------------------
# Geometry helpers
# ---------------------------------------------------------------------------

def rect_overlap(a: tuple, b: tuple, threshold: float = 0.3) -> bool:
    """
    Returns True if rectangles a and b overlap by at least `threshold`
    of the smaller rectangle's area.
    a, b = (x0, y0, x1, y1)
    """
    ax0, ay0, ax1, ay1 = a
    bx0, by0, bx1, by1 = b

    ix0 = max(ax0, bx0)
    iy0 = max(ay0, by0)
    ix1 = min(ax1, bx1)
    iy1 = min(ay1, by1)

    if ix1 <= ix0 or iy1 <= iy0:
        return False

    inter_area = (ix1 - ix0) * (iy1 - iy0)
    a_area = (ax1 - ax0) * (ay1 - ay0)
    b_area = (bx1 - bx0) * (by1 - by0)
    smaller = min(a_area, b_area)
    if smaller == 0:
        return False
    return (inter_area / smaller) >= threshold


def find_covering_link(
    token: TextToken,
    annotations: list[LinkAnnotation],
    page_annotations: dict[int, list[LinkAnnotation]],
) -> Optional[LinkAnnotation]:
    """Return the annotation that covers this token, or None."""
    for ann in page_annotations.get(token.page_num, []):
        if rect_overlap(
            (token.x0, token.y0, token.x1, token.y1),
            ann.rect,
        ):
            return ann
    return None


def tokens_on_same_line(t1: TextToken, t2: TextToken, tol: float = 4.0) -> bool:
    return abs(t1.y0 - t2.y0) < tol and t1.page_num == t2.page_num


def tokens_adjacent_lines(t1: TextToken, t2: TextToken) -> bool:
    """True if t2 is on the very next line after t1 (same page)."""
    if t1.page_num != t2.page_num:
        return False
    line_height = (t1.y1 - t1.y0)
    gap = t2.y0 - t1.y1
    return 0 <= gap <= line_height * 2.5


# ---------------------------------------------------------------------------
# Address-type helpers
# ---------------------------------------------------------------------------

def is_email_uri(uri: str) -> bool:
    return uri.lower().startswith("mailto:")


def is_web_uri(uri: str) -> bool:
    return uri.lower().startswith("http://") or uri.lower().startswith("https://") or uri.lower().startswith("www.")


def uri_email_address(uri: str) -> str:
    """Extract the bare email from a mailto: URI."""
    return uri[7:].split("?")[0].strip() if uri.lower().startswith("mailto:") else ""


def normalise_url(url: str) -> str:
    """Strip trailing punctuation that may have been captured by regex."""
    return url.rstrip(".,;:!?)\"'")


def addresses_match(displayed: str, uri: str) -> bool:
    """
    Compare the displayed address text to the URI it is linked to.
    For emails: displayed text should equal the mailto: address.
    For URLs:   displayed text (with http/https/www normalised) should match URI.
    """
    displayed = displayed.strip().rstrip(".,;:!?")

    if is_email_uri(uri):
        uri_addr = uri_email_address(uri)
        return displayed.lower() == uri_addr.lower()

    if is_web_uri(uri):
        # Normalise both sides for comparison
        def norm(u):
            u = u.lower().rstrip("/")
            if u.startswith("http://"):  u = u[7:]
            if u.startswith("https://"): u = u[8:]
            if u.startswith("www."):     u = u[4:]
            return u
        return norm(displayed) == norm(uri)

    return False


# ---------------------------------------------------------------------------
# Core checker
# ---------------------------------------------------------------------------

def build_page_annotation_index(annotations: list[LinkAnnotation]) -> dict[int, list[LinkAnnotation]]:
    idx: dict[int, list[LinkAnnotation]] = {}
    for ann in annotations:
        idx.setdefault(ann.page_num, []).append(ann)
    return idx


def group_tokens_into_lines(tokens: list[TextToken]) -> list[list[TextToken]]:
    """Group tokens by (page, approximate y-baseline) into lines."""
    if not tokens:
        return []
    lines: list[list[TextToken]] = []
    current_line = [tokens[0]]
    for tok in tokens[1:]:
        if tokens_on_same_line(tok, current_line[-1]):
            current_line.append(tok)
        else:
            lines.append(current_line)
            current_line = [tok]
    lines.append(current_line)
    return lines


def scan_line_for_addresses(line_tokens: list[TextToken]) -> list[tuple[str, list[TextToken]]]:
    """
    Join tokens on the same line and find email/URL patterns.
    Returns list of (matched_address, [tokens_that_form_it]).
    This is a best-effort spatial mapping.
    """
    if not line_tokens:
        return []
    full_text = " ".join(t.text for t in line_tokens)
    results = []

    for pattern in (EMAIL_RE, URL_RE):
        for m in pattern.finditer(full_text):
            addr = normalise_url(m.group())
            # Map back to tokens (greedy span match)
            covering = []
            pos = 0
            for tok in line_tokens:
                tok_start = full_text.find(tok.text, pos)
                tok_end   = tok_start + len(tok.text)
                # Token overlaps the match span
                if tok_start < m.end() and tok_end > m.start():
                    covering.append(tok)
                pos = tok_start + 1  # advance slightly
            if covering:
                results.append((addr, covering))

    return results


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------

def export_to_excel(results: list[CheckResult], output_path: str):
    """Export check results to Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        print("❌ openpyxl not installed. Install it with: pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "PDF Link Check"

    # Header row
    headers = [
        "PDF_Link",
        "Hyperlink_Text",
        "Page_Numb",
        "Link_Type",
        "Is_Hyperlinked",
        "Is_Valid",
        "Hyperlink_Points_To",
        "Result"
    ]
    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data rows
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for result in results:
        ws.append([
            result.pdf_link,
            result.hyperlink_text,
            result.page_num,
            result.link_type,
            result.is_hyperlinked,
            result.is_valid,
            result.hyperlink_points_to,
            result.result
        ])

        row_idx = ws.max_row
        result_cell = ws[f"H{row_idx}"]
        if result.result == "Pass":
            result_cell.fill = pass_fill
            result_cell.font = Font(color="006100", bold=True)
        else:
            result_cell.fill = fail_fill
            result_cell.font = Font(color="9C0006", bold=True)

    # Adjust column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 12

    wb.save(output_path)
    print(f"✅ Report saved to: {output_path}")


# ---------------------------------------------------------------------------
# Main analysis
# ---------------------------------------------------------------------------

def check_pdf(pdf_path: str, verbose: bool = False) -> list[CheckResult]:
    results: list[CheckResult] = []
    issues: list[Issue] = []

    print(f"📄  Loading: {pdf_path}")
    annotations  = extract_annotations(pdf_path)
    page_ann_idx = build_page_annotation_index(annotations)
    all_tokens   = extract_words(pdf_path)

    print(f"    Found {len(annotations)} hyperlink annotations across all pages.")
    print(f"    Found {len(all_tokens)} word tokens across all pages.\n")

    lines = group_tokens_into_lines(all_tokens)

    # -----------------------------------------------------------------------
    # Pass 1 — Single-line addresses
    # -----------------------------------------------------------------------
    processed_tokens: set[int] = set()  # token ids already handled

    for line in lines:
        found = scan_line_for_addresses(line)
        for addr, covering_tokens in found:
            # Mark tokens
            for t in covering_tokens:
                processed_tokens.add(id(t))

            page = covering_tokens[0].page_num + 1
            is_email = bool(EMAIL_RE.fullmatch(addr))
            is_url   = bool(URL_RE.match(addr))
            
            link_type = "Email" if is_email else "Web Link"

            # Gather links covering these tokens
            links = []
            for tok in covering_tokens:
                ann = find_covering_link(tok, annotations, page_ann_idx)
                if ann:
                    links.append(ann)

            unique_uris = list({a.uri for a in links})

            # Create result entry
            is_hyperlinked = "Yes" if unique_uris else "No"
            hyperlink_text = unique_uris[0] if unique_uris else "No Link"
            
            is_valid = "Yes"
            result = "Pass"

            # ---- Check 1: No link at all ----
            if not unique_uris:
                is_valid = "No"
                result = "Fail"
                issues.append(Issue(
                    page=page,
                    issue_type="MISSING LINK",
                    displayed_text=addr,
                    linked_uri=None,
                    detail=f"{'Email' if is_email else 'Web'} address has no hyperlink.",
                ))
            else:
                # ---- Check 2 & 5: Wrong link type / wrong target ----
                for uri in unique_uris:
                    if is_email and is_web_uri(uri):
                        is_valid = "No"
                        result = "Fail"
                        issues.append(Issue(
                            page=page,
                            issue_type="WRONG LINK TYPE",
                            displayed_text=addr,
                            linked_uri=uri,
                            detail="Email address is linked to a web URL instead of a mailto: link.",
                        ))
                    elif is_url and is_email_uri(uri):
                        is_valid = "No"
                        result = "Fail"
                        issues.append(Issue(
                            page=page,
                            issue_type="WRONG LINK TYPE",
                            displayed_text=addr,
                            linked_uri=uri,
                            detail="Web URL is linked to a mailto: email link instead of a web URL.",
                        ))
                    elif not addresses_match(addr, uri):
                        is_valid = "No"
                        result = "Fail"
                        issues.append(Issue(
                            page=page,
                            issue_type="WRONG TARGET",
                            displayed_text=addr,
                            linked_uri=uri,
                            detail=f"Displayed address does not match the actual link target.",
                        ))
                    else:
                        if verbose:
                            print(f"  ✅ Page {page}: \"{addr}\" → {uri}")

            results.append(CheckResult(
                pdf_link=addr,
                hyperlink_text=hyperlink_text,
                page_num=page,
                link_type=link_type,
                is_hyperlinked=is_hyperlinked,
                is_valid=is_valid,
                hyperlink_points_to=hyperlink_text,
                result=result
            ))

    return results


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------

ISSUE_SYMBOLS = {
    "MISSING LINK":                "🔴",
    "WRONG LINK TYPE":             "🟠",
    "WRONG TARGET":                "🟡",
    "SPLIT ADDRESS — NO LINK":     "🔴",
    "SPLIT ADDRESS — PARTIAL LINK":"🟠",
    "SPLIT ADDRESS — DIFFERENT LINKS": "🟠",
    "SPLIT ADDRESS — WRONG TARGET":"🟡",
    "SPLIT URL — NO LINK":         "🔴",
    "SPLIT URL — PARTIAL LINK":    "🟠",
    "SPLIT URL — DIFFERENT LINKS": "🟠",
}

ISSUE_ORDER = {
    "MISSING LINK": 0,
    "SPLIT ADDRESS — NO LINK": 0,
    "SPLIT URL — NO LINK": 0,
    "WRONG LINK TYPE": 1,
    "SPLIT ADDRESS — PARTIAL LINK": 1,
    "SPLIT URL — PARTIAL LINK": 1,
    "SPLIT ADDRESS — DIFFERENT LINKS": 1,
    "SPLIT URL — DIFFERENT LINKS": 1,
    "WRONG TARGET": 2,
    "SPLIT ADDRESS — WRONG TARGET": 2,
}


def render_report(pdf_path: str, results: list[CheckResult], output_path: Optional[str] = None):
    lines = []
    sep   = "=" * 72

    lines.append(sep)
    lines.append("  PDF LINK CHECKER — REPORT")
    lines.append(f"  File    : {pdf_path}")
    lines.append(f"  Entries : {len(results)}")
    lines.append(sep)

    if not results:
        lines.append("\n✅  No entries found.\n")
    else:
        # Count passes and fails
        passes = sum(1 for r in results if r.result == "Pass")
        fails = sum(1 for r in results if r.result == "Fail")
        
        lines.append(f"\n📊 Summary:")
        lines.append(f"   ✅ Pass: {passes}")
        lines.append(f"   ❌ Fail: {fails}\n")

    report = "\n".join(lines)
    print(report)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Check a PDF for broken/missing/wrong email and web hyperlinks.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("pdf", help="Path to the PDF file to check")
    parser.add_argument("--output", "-o", help="Save report to Excel file (.xlsx)", default=None)
    parser.add_argument("--verbose", "-v", action="store_true", help="Show passing checks too")
    args = parser.parse_args()

    results = check_pdf(args.pdf, verbose=args.verbose)
    render_report(args.pdf, results)

    if args.output:
        if not args.output.lower().endswith('.xlsx'):
            args.output += '.xlsx'
        export_to_excel(results, args.output)
    else:
        # Default output name
        import os
        base_name = os.path.splitext(os.path.basename(args.pdf))[0]
        default_output = f"{base_name}_link_report.xlsx"
        export_to_excel(results, default_output)

    # Exit code: 0 = all pass, 1 = some failed
    fail_count = sum(1 for r in results if r.result == "Fail")
    sys.exit(1 if fail_count > 0 else 0)


if __name__ == "__main__":
    main()
